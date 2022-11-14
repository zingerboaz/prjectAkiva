from __future__ import unicode_literals
import json

from rest_framework import status
from rest_framework.decorators import api_view
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

from ba_main.models import InstitutionDetails
from ba_main.models.institution import Institution
from ba_main.serializers.institution import SingleInstitutionSerializer, InstitutionSerializer, ImageSerializer
from django.utils.translation import ugettext_lazy as _
from rest_framework import viewsets, permissions

# @api_view(['GET'])
# @permission_classes([IsAuthenticated])
# def get_institution_list_by_user(request):
#     current_user = User.objects.get(email=request.user)
#
#     if current_user.institution_id:
#         return get_one_institution(current_user.institution_id)
#     else:
#         return get_all_institutions()
from ba_main.serializers.institution_yearly_detail import InstitutionYearlyDetailSerializer
from ba_main.serializers.story_and_strength import StoryAndStrengthSerializer
from data_pulling.fetching import insert_indexes_to_db_from_json
from proj.utils import Utils
from rest_framework.views import APIView

from ba_main.models.user import User

from ba_main import models
from ba_main.models.institutionprogram import InstitutionProgram
from ba_main.serializers.institution_program import ProgramSerializer


@api_view(['GET'])
# @permission_classes([IsAuthenticated])
def get_institution_list(request):
    return get_all_institutions()


def get_all_institutions():
    institutions = Institution.objects.all()
    serializer = InstitutionSerializer(institutions, many=True)
    return Response(serializer.data)


def get_one_institution(institution_id):
    try:
        institution = Institution.objects.get(id=institution_id)
        serializer = SingleInstitutionSerializer(institution)
        return Response(serializer.data)
    except Institution.DoesNotExist:
        return Response({_('Institution not found')},
                        status=status.HTTP_404_NOT_FOUND)


@api_view(['PUT'])
def update_story_and_strength(request):
    # filter_keys = ['institution_id', 'year', 'story', 'strength']
    # story_and_strength = {k: v for (k, v) in request.data.items() if k in filter_keys}

    story_and_strength_serializer = StoryAndStrengthSerializer(data=request.data)

    if not story_and_strength_serializer.is_valid():
        return Utils.create_invalid_data_error_response(
            request.data,
            story_and_strength_serializer.errors)
    request_params = story_and_strength_serializer.data

    try:
        institution_details: InstitutionDetails = InstitutionDetails.objects.get(
            institution__id=request_params.get('institution_id'),
            year__name=request_params.get('year')
        )
        institution_details.institution_story = request_params.get('story')
        institution_details.institution_strength = request_params.get('strength')
        institution_details.institution_challenge = request_params.get('challenge')
        institution_details.save()
        institution_details_serializer = InstitutionYearlyDetailSerializer(institution_details)

        plans_keys = ['plan_1', 'plan_2', 'plan_3']
        for plan_key_name in plans_keys:
            if request.data.get(plan_key_name) and request.data.get(plan_key_name)['name_of_program']:
                program_serializer = ProgramSerializer(data=request.data.get(plan_key_name))
                if not program_serializer.is_valid():
                    return Utils.create_invalid_data_error_response(
                        request.data,
                        program_serializer.errors)
                program_data_object = program_serializer.data
                program_instance_args = {
                    'program_type': plan_key_name,
                    'name': program_data_object.get('name_of_program'),
                    'purpose': program_data_object.get('goal_of_program'),
                    'target_audience': program_data_object.get('target_audience'),
                    'regularity': program_data_object.get('regularity'),
                    'success_factors': program_data_object.get('success_factors'),
                    'institution_details': institution_details
                }
                print('good')
                # print(program_data_object.get('institution_id'))
                InstitutionProgram.objects.update_or_create(institution_details__id=institution_details.id ,program_type=plan_key_name, defaults=program_instance_args)
            else:
                print('else')
                try:
                    InstitutionProgram.objects.filter(institution_details__id=institution_details.id,
                                      program_type=plan_key_name).delete()
                except:
                    pass

        return Utils.create_instance_updated_response(
            'Institution Details',
            institution_details_serializer.data, )

    except InstitutionDetails.DoesNotExist:
        return Utils.create_instance_not_found_response('Institution Details')


@api_view(['GET', 'POST'])
def institution_image(request, institution_id):
    institution = Institution.objects.get(id=institution_id)
    if institution:
        if request.method == 'POST':
            institution.image = request.FILES["file"]
            institution.save()
        serializer = ImageSerializer(institution)
        print(serializer.data)
        return Response(serializer.data, status=status.HTTP_200_OK)
    else:
        return Response("institution not found", status=status.HTTP_400_BAD_REQUEST)



@api_view(['POST'])
def get_indices_torani(request):

    try:
        from django.core.validators import FileExtensionValidator
        from django import forms

        class UploadFile(forms.Form):
            data_file = forms.FileField(validators=[FileExtensionValidator(['xlsx'])])

        form = UploadFile(request.POST, request.FILES)
        if not form.is_valid():
            return Response("failed to read file", status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        from openpyxl import load_workbook
        from zipfile import ZipFile
        work_book = load_workbook(filename=request.FILES['data_file'].file)
        work_sheet = work_book.active

        groups_obj = {}
        groups = []
        classes = []
        indexes = []
        groups_obj['groups'] = groups

        semel_mosad = int(request.data.get('semel_mosad'))
        for row in work_sheet.iter_rows(min_row=2):
            try:
                row = list(row)
                institution_semel = row.__getitem__(0).value
                if institution_semel is None:
                    break

                institution_semel = int(institution_semel)
                if institution_semel != semel_mosad:
                    continue

                id = row.__getitem__(1).value
                year = int(row.__getitem__(2).value)
                HByear = row.__getitem__(3).value
                name_of_category = row.__getitem__(4).value
                indice_institution = row.__getitem__(5).value
                indice_similar_institution = row.__getitem__(6).value
                indice_national_average = row.__getitem__(7).value
                print('for2')
                index = get_index_by_id(id, indexes)
                compare_values: [] = fill_compare_values(indice_institution, indice_similar_institution,
                                                         indice_national_average)
                obj: dict = {
                    'institution_semel': semel_mosad,
                    'Id': id,
                    'Year': year,
                    'HebYear': HByear,
                    'Name':name_of_category,
                    'name_of_category': name_of_category,
                    'CompareValuesModel': compare_values,
                }
                if index:
                    index['History'].append(obj)
                else:
                    history = obj.copy()
                    obj['History'] = [history]
                    indexes.append(obj)
            except Exception:
                print('excetion for get_indices_torani')
                pass

        classes.append({'Indexes': indexes})
        groups.append({'Name':'בגרות','Classes': classes})
        print(groups_obj)
        # groups_json = json.dumps(groups_obj)
        insert_indexes_to_db_from_json(semel_mosad, groups_obj)
        return Response("success", status=status.HTTP_200_OK)
    except Exception as e:
        return Response("failed to read file", status=status.HTTP_500_INTERNAL_SERVER_ERROR)


def get_index_by_id(id: str, indexes: []):
    for index in indexes:
        if index['Id'] == id:
            return index
    return None


def fill_compare_values(indice_institution, indice_similar_institution, indice_national_average):
    return [{
        'Id': 0,
        'Name': 'מוסד נבחר',
        'Value': indice_institution
    }, {
        'Id': 1,
        'Name': 'מוסדות דומים',
        'Value': indice_similar_institution
    }, {
        'Id': 2,
        'Name': 'ממוצע ארצי ',
        'Value': indice_national_average
    }]


# convert xls feil into xlsx feil
# def cvt_xls_to_xlsx(src_file_path, dst_file_path):
#     import xlrd
#     from openpyxl.workbook import Workbook
#
#     book_xls = xlrd.open_workbook(src_file_path)
#     book_xlsx = Workbook()
#
#     sheet_names = book_xls.sheet_names()
#     for sheet_index, sheet_name in enumerate(sheet_names):
#         sheet_xls = book_xls.sheet_by_name(sheet_name)
#         if sheet_index == 0:
#             sheet_xlsx = book_xlsx.active
#             sheet_xlsx.title = sheet_name
#         else:
#             sheet_xlsx = book_xlsx.create_sheet(title=sheet_name)
#
#         for row in range(0, sheet_xls.nrows):
#             for col in range(0, sheet_xls.ncols):
#                 sheet_xlsx.cell(row = row+1 , column = col+1).value = sheet_xls.cell_value(row, col)
#
#     book_xlsx.save(dst_file_path)

@api_view(['POST'])
def institutions_base_details(request):
    try:
        from django.core.validators import FileExtensionValidator
        from django import forms
        class UploadFileForm(forms.Form):
            data_file = forms.FileField(validators=[FileExtensionValidator(['xlsx'])])

        form = UploadFileForm(request.POST, request.FILES)
        if not form.is_valid():
            return
        from openpyxl import load_workbook
        work_book = load_workbook(filename=request.FILES['data_file'].file)
        work_sheet = work_book.active

        for row in work_sheet.iter_rows(min_row=2):
            institution_semel = 0
            try:
                row = list(row)
                institution_semel = int(row.__getitem__(0).value)
                institution_name = row.__getitem__(1).value
                area_manager_first_name = row.__getitem__(2).value
                area_manager_last_name = row.__getitem__(3).value
                area_manager_email = row.__getitem__(4).value
                manager_first_name = row.__getitem__(5).value
                manager_last_name = row.__getitem__(6).value
                manager_email = row.__getitem__(7).value
                institution_inspector = row.__getitem__(8).value
                is_boarding = False
                if row.__getitem__(9).value:
                    is_boarding = row.__getitem__(9).value
                full_ownership = row.__getitem__(10).value

                print('institution_semel', institution_semel)
                institution, created = Institution.objects.get_or_create(semel=institution_semel)
                institution.name = institution_name
                institution.inspector = institution_inspector
                institution.is_boarding = is_boarding
                institution.full_ownership = full_ownership

                try:
                    manager, created = update_or_create_user(manager_email, manager_first_name, manager_last_name)
                    if created:
                        manager.make_manager()
                    manager.save()
                    institution.manager = manager
                except Exception:
                    print("failed to create or update institution manager for: " + f'{institution_semel}'
                          + "with email: " + manager_email)

                try:
                    area_manager, created = update_or_create_user(area_manager_email
                                                         , area_manager_first_name, area_manager_last_name)
                    if created:
                        area_manager.make_area_manager()
                    area_manager.save()
                    print(User.objects.get_or_create(email=area_manager_email))
                    institution.area_manager = area_manager
                except Exception:
                    print("failed to create or update area manager for: " + f'{institution_semel}'
                          + "with email: " + area_manager_email)

                try:
                    institution.save()
                    print(f'{institution.semel}' + "success")
                except Exception:
                    print("failed to create or update institution: " + f'{institution_semel}' + " but manager and area manager success")

            except Exception:
                print("failed to create or update institution: " + f'{institution_semel}')
            
        # delete the institutions that not in the work_book
        for i, institution in enumerate(Institution.objects.all()):
            found = 0
            # print(f"-{int(institution.semel)}")
            for row in work_sheet.iter_rows(min_row=2):
                if int(institution.semel) == row.__getitem__(0).value:
                    # print(f"  {institution.semel} == {row.__getitem__(0).value}")
                    found = 1
            if not found:
                print(f"{institution.semel} deleted.")
                institution.delete()
            else:
                print(f"{institution.semel} found.")

        return Response("success", status=status.HTTP_200_OK)

    except Exception:
        return Response("failed to read file", status=status.HTTP_500_INTERNAL_SERVER_ERROR)


def update_or_create_user(email, first_name, last_name):
    manager, m_created = User.objects.get_or_create(
        email=email)

    manager.first_name = first_name
    manager.last_name = last_name
    manager.user_name = first_name
    if m_created:
        manager.password = first_name
    return manager, m_created
