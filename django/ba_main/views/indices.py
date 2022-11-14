import math

from django import forms
from django.http import HttpResponseRedirect
from rest_framework.decorators import api_view
from rest_framework.request import Request
from rest_framework.response import Response

from ba_main.models import YearlyIndice, InstitutionDetails, YearlyGoal, Year, IndiceType
from ba_main.models.group_name import GroupName
from ba_main.models.indice import InstitutionIndice
from ba_main.models.indice_group import IndiceGroup
from ba_main.models.indice_subject import IndiceSubject
from ba_main.models.institution import Institution
from ba_main.models.subject_name import SubjectName
from ba_main.serializers.conclusion import UpdateConclusionRequestSerializer
from ba_main.serializers.yearly_indice import YearlyIndiceRequestSerializer, \
    IndiceGroupSerializer, AllYearsEducationIndiceSerializer, IndiceSubjectSerializer, BrieflyIndiceSerializer, \
    DataTypeAvailableYears, YearModelSerializer
from proj.utils import Utils


@api_view(['POST'])
def get_data_type_available_years(request: Request):
    request_data_serializer = DataTypeAvailableYears(data=request.data)
    if not request_data_serializer.is_valid():
        return Utils.create_invalid_data_error_response(request.data, request_data_serializer.errors)
    request_data = request_data_serializer.data
    institution_id, data_type = request_data.get('institution_id'), request_data.get('data_type')

    if data_type in ['SOCIAL', 'EDUCATION', 'PARTICIPATION', 'STAFF']:
        institution_type_yearly_indices = YearlyIndice.objects\
            .filter(institution__id=institution_id, type__name=data_type)\
            .order_by('-year__name')
        indices_years_list = []
        for yearly_indices_item in list(institution_type_yearly_indices):
            if yearly_indices_item.groups and len(list(yearly_indices_item.groups.all())):
                serialized_year_data = YearModelSerializer(yearly_indices_item.year)
                indices_years_list.append(serialized_year_data.data)

        return Response({'indices_type': data_type, 'available_years': indices_years_list})

    elif data_type == 'INSTITUTION_DETAILS':
        available_institution_details_years = []
        all_years_institution_details = InstitutionDetails.objects\
            .filter(institution__id=institution_id)\
            .order_by('-year__name')
        for inst_details_by_year in list(all_years_institution_details):
            serialized_year_data = YearModelSerializer(inst_details_by_year.year)
            available_institution_details_years.append(serialized_year_data.data)
        return Response({'indices_type': data_type, 'available_years': available_institution_details_years})

    elif data_type == 'YEARLY_GOALS':
        available_institution_goals_years = []
        all_years_institution_goals = YearlyGoal.objects\
            .filter(institution__id=institution_id)\
            .order_by('-year__name')
        for yearly_goal in list(all_years_institution_goals):
            if not yearly_goal.year in available_institution_goals_years:
                available_institution_goals_years.append(yearly_goal.year)

        available_years_data_serializer = YearModelSerializer(available_institution_goals_years, many=True)
        return Response({'indices_type': data_type, 'available_years': available_years_data_serializer.data})

    else:
        return Utils.create_instance_not_found_response(data_type)


@api_view(['POST'])
def get_yearly_education_indice(request):
    request_data = fetch_indice_args_from_request(request)
    yearly_education_indice = IndiceGroup.objects.filter(
        yearly_indice__year__name=request_data.get('year'),
        yearly_indice__institution=request_data.get('institution_id'),
        yearly_indice__type__name='EDUCATION'
    )

    yearly_indice_response_serializer = IndiceGroupSerializer(
        yearly_education_indice,
        many=True,
        context={'selected_institutions': request_data.get('selected_institutions')})
    # print(yearly_indice_response_serializer.data)
    return Response(yearly_indice_response_serializer.data)


@api_view(['POST'])
def get_all_years_education_indice(request):
    request_data_serializer = AllYearsEducationIndiceSerializer(data=request.data)
    if not request_data_serializer.is_valid():
        return Response('Request Data invalid!')
    request_data = request_data_serializer.data
    all_years_education_indice = InstitutionIndice.objects.filter(
        subject__indice_group__yearly_indice__institution=request_data.get('institution_id'),
        subject__indice_group__yearly_indice__type__name='EDUCATION',
        subject__name__id=request_data.get('subject_name_id')

    ).order_by('subject__indice_group__yearly_indice__year__name')
    subject = IndiceSubject.objects.get(name__id=request_data.get('subject_name_id'))
    all_years_education_indice_response = BrieflyIndiceSerializer(all_years_education_indice, many=True)
    return Response(
        {
            'subject': IndiceSubjectSerializer(subject).data['name'],
            'indices': all_years_education_indice_response.data
        })


def fetch_indice_args_from_request(request: Request):
    yealy_indice_serializer = YearlyIndiceRequestSerializer(data=request.data)
    if yealy_indice_serializer.is_valid():
        request_data = yealy_indice_serializer.data
        return request_data
    else:
        return None


@api_view(['POST'])
def get_yearly_participation_indice(request):
    request_data = fetch_indice_args_from_request(request)
    yearly_participation_indice = IndiceGroup.objects.filter(
        yearly_indice__year__name=request_data.get('year'),
        yearly_indice__institution=request_data.get('institution_id'),
        yearly_indice__type__name='PARTICIPATION'
    )
    yearly_indice_response = IndiceGroupSerializer(
        yearly_participation_indice,
        many=True,
        context={
            'selected_institutions': request_data.get('selected_institutions'),
            'own_only': True
        }
    )
    return Response(yearly_indice_response.data)


@api_view(['POST'])
def get_yearly_staff_indice(request):
    request_data = fetch_indice_args_from_request(request)
    yearly_staff_indice = IndiceGroup.objects.filter(
        yearly_indice__year__name=request_data.get('year'),
        yearly_indice__institution=request_data.get('institution_id'),
        yearly_indice__type__name='STAFF'
    )
    yearly_indice_response = IndiceGroupSerializer(
        yearly_staff_indice,
        many=True,
        context={'selected_institutions': request_data.get('selected_institutions')}
    )

    return Response(yearly_indice_response.data)


@api_view(['POST'])
def get_yearly_social_indices(request):
    request_data = fetch_indice_args_from_request(request)
    yearly_social_indices = IndiceGroup.objects.filter(
        yearly_indice__year__name=request_data.get('year'),
        yearly_indice__institution=request_data.get('institution_id'),
        yearly_indice__type__name='SOCIAL'
    )
    yearly_indice_response_serializer = IndiceGroupSerializer(
        yearly_social_indices,
        many=True,
        context={'selected_institutions': request_data.get('selected_institutions')})
    return Response(yearly_indice_response_serializer.data)


@api_view(['POST'])
def get_yearly_indices_conclusion(request):
    request_serializer = UpdateConclusionRequestSerializer(data=request.data)
    request_valid = request_serializer.is_valid()
    if not request_valid:
        return Response({'error': 'request data not valid'})
    request_data = request_serializer.data
    yearly_indice = YearlyIndice.objects.get(id=request_data.get('yearly_indice_id'))
    return Response(UpdateConclusionRequestSerializer(data=yearly_indice))




# @api_view(['POST'])
# def get_indices_torani(request):
#     print(request.body)
#     return Response('fgfgfg')






@api_view(['POST'])
def uploading_participation_data(request):
    from django.core.validators import FileExtensionValidator
    class UploadFileForm(forms.Form):
        file = forms.FileField(validators=[FileExtensionValidator(['xlsx'])])
        year = forms.CharField()
        participation_type = forms.CharField()

    form = UploadFileForm(request.POST, request.FILES)
    if not form.is_valid():
        return
    from openpyxl import load_workbook
    work_book = load_workbook(filename=request.FILES['file'].file)
    # print('year', form.data.get('year'))
    # print('type', form.data.get('participation_type'))
    participation_year, participation_year_created = Year.objects.get_or_create(name=form.data.get('year'))
    participation_type = form.data.get('participation_type')
    group_name, group_name_created = GroupName.objects\
        .get_or_create(title=participation_type, defaults={'title': participation_type})

    yearly_indices_type, yit_created = IndiceType.objects\
        .get_or_create(name='PARTICIPATION', defaults={'name':'PARTICIPATION'})

    work_sheet = work_book.active
    participation_names_row = work_sheet[1][2:]
    participation_names = []
    participation_count = 0
    visited_participation_count = 0

    for name in participation_names_row:
        if name.value:
            participation_names.append(name.value)
        else:
            break
    for row in work_sheet.iter_rows(min_row=2):
        row = list(row)
        try:
            institution_semel = math.floor(row.pop(0).value)
        except:
            pass
        print(institution_semel)
        if institution_semel is None:
            break
        institution = None
        try:
            # print('institution_semel', institution_semel)
            institution = Institution.objects.get(semel=institution_semel)

            yearly_indices_args = { 'type': yearly_indices_type,
                                    'institution': institution,
                                    'year': participation_year
                                    }

            yearly_indices, yi_created = YearlyIndice.objects\
                .update_or_create(
                year__id=participation_year.id,
                institution__id=institution.id,
                type__id=yearly_indices_type.id,
                defaults=yearly_indices_args
            )
            group_args = {'name': group_name, 'yearly_indice': yearly_indices}
            group, group_created = IndiceGroup.objects \
                .update_or_create(name__id=group_name.id, yearly_indice__id=yearly_indices.id, defaults=group_args)
            group.yearly_indice = yearly_indices
            group.save()

            for name_index, name in enumerate(participation_names):
                cell_value = row[name_index + 1].value
                if cell_value != None:
                    name = name.strip()
                    subject_name_args = {'title': name }
                    subject_name, sn_created = SubjectName.objects\
                        .update_or_create(title=name, defaults=subject_name_args)
                    subject_args ={ 'name': subject_name, 'indice_group': group}
                    subject, subject_created = IndiceSubject.objects\
                        .update_or_create(name__id=subject_name.id, indice_group__id=group.id, defaults=subject_args)
                    value = 0
                    if cell_value:
                        value = 1
                        visited_participation_count += 1
                    indice_args = {'subject': subject, 'value': value, 'chativa': None }
                    indices, i_created = InstitutionIndice.objects\
                        .update_or_create(subject__id=subject.id, defaults=indice_args)
                    participation_count += 1
                    print(f'{indices.subject.name.title} - {indices.value}')
                else:
                    print(cell_value)


# CALCULATING VISITED PARTICIPATION PERCENTAGE
            visited_participation = 0
            if participation_count:
                visited_participation = (( visited_participation_count / participation_count) * 100)

            visited_participation_group_name, vp_grn_created = GroupName.objects.update_or_create(
                title='VISITED PARTICIPATION', defaults={'title': 'VISITED PARTICIPATION'})
            visited_participation_group, vp_group_created = IndiceGroup.objects.update_or_create(
                name__id=visited_participation_group_name.id,
                yearly_indice__id=yearly_indices.id,
                defaults={'name':visited_participation_group_name, 'yearly_indice': yearly_indices }
            )
            vp_subject_name, vp_sbj_name_created = SubjectName.objects.update_or_create(
                title=participation_type, defaults={'title': participation_type}
            )
            vp_subject_args = {'name': vp_subject_name, 'indice_group': visited_participation_group}
            vp_subject, sbj_created = IndiceSubject.objects.update_or_create(
                name__id=vp_subject_name.id,
                indice_group__id=visited_participation_group.id,
                defaults=vp_subject_args
            )
            vp_indices_args = {'subject': vp_subject, 'value': visited_participation, 'chativa': None}
            visited_participation_indices, vsp_indices_created = InstitutionIndice.objects.update_or_create(
                subject__id=vp_subject.id, defaults=vp_indices_args
            )


        except Institution.DoesNotExist as e:
            print(e)
            continue
        except Exception as e:
            print(e)
            continue
    return HttpResponseRedirect(request.META.get('HTTP_REFERER'))

